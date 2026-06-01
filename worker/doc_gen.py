"""Excel / PDF 服务端生成 — 子项目③。

设计: docs/superpowers/specs/2026-05-30-excel-pdf-多模态生成-design.md
硬指标: PDF 中文不能出豆腐块 → reportlab 内置 STSong-Light (Adobe-GB1) CID 字体。
纯函数(build_excel/build_pdf 返回 bytes, 不碰网络/存储), 可单测; worker 负责上传。
"""
from __future__ import annotations

import io
import uuid

# ── 中文字体: reportlab 内置 CID 字体, 自带简体中文, 无需打包字体文件 ──────────
_CN_FONT = "STSong-Light"
_font_ready = False


def _ensure_cn_font():
    """注册中文 CID 字体(只注册一次)。"""
    global _font_ready
    if _font_ready:
        return
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    pdfmetrics.registerFont(UnicodeCIDFont(_CN_FONT))
    _font_ready = True


# ── 文件名 / 存储路径 ────────────────────────────────────────────────────────

def storage_path(filename: str, ext: str) -> str:
    """中文显示名 → ASCII storage 路径 (中文路径 Supabase Storage 返回 400, 踩过坑)。
    用 uuid 保唯一; 中文名只用于 attachment.name 展示。"""
    return f"gen/{uuid.uuid4().hex}.{ext}"


# ── Excel ────────────────────────────────────────────────────────────────────

def build_excel(spec: dict) -> bytes:
    """结构化数据 → .xlsx bytes。
    spec: {filename, sheet_name?, title?, headers[], rows[][], summary?[][] }
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = (spec.get("sheet_name") or "Sheet1")[:31]

    bold = Font(bold=True)
    r = 1
    title = spec.get("title")
    if title:
        ws.cell(row=r, column=1, value=title).font = Font(bold=True, size=14)
        r += 2

    headers = spec.get("headers") or []
    if headers:
        for c, h in enumerate(headers, start=1):
            ws.cell(row=r, column=c, value=h).font = bold
        r += 1

    for row in (spec.get("rows") or []):
        for c, v in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=v)
        r += 1

    for row in (spec.get("summary") or []):
        for c, v in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = bold
        r += 1

    # 自适应列宽(粗略)
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max(width + 2, 10), 50)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── PDF ──────────────────────────────────────────────────────────────────────

def _build_pdf_bytes(spec: dict, scale: float = 1.0) -> bytes:
    """渲染一次 PDF。scale<1 时等比缩小字号/行距/边距/留白/表格 padding,用于压页。"""
    _ensure_cn_font()
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    )

    styles = getSampleStyleSheet()
    body = ParagraphStyle("CNBody", parent=styles["Normal"],
                          fontName=_CN_FONT, fontSize=11 * scale, leading=18 * scale)
    h1 = ParagraphStyle("CNTitle", parent=styles["Title"],
                        fontName=_CN_FONT, fontSize=18 * scale, leading=24 * scale)

    story = []
    title = spec.get("title")
    if title:
        story.append(Paragraph(title, h1))
        story.append(Spacer(1, 6 * mm * scale))

    for blk in (spec.get("blocks") or []):
        btype = blk.get("type")
        if btype == "paragraph":
            story.append(Paragraph(str(blk.get("text", "")), body))
            story.append(Spacer(1, 3 * mm * scale))
        elif btype == "table":
            headers = blk.get("headers") or []
            rows = blk.get("rows") or []
            data = ([headers] if headers else []) + [list(map(str, r)) for r in rows]
            if not data:
                continue
            # 用 Paragraph 包裹单元格以支持中文字体
            cell_style = ParagraphStyle("CNCell", parent=body,
                                        fontSize=10 * scale, leading=14 * scale)
            wrapped = [[Paragraph(str(c), cell_style) for c in row] for row in data]
            tbl = Table(wrapped, hAlign="LEFT")
            pad = max(1, round(4 * scale))
            ts = [
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F0F0F0")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), pad),
                ("BOTTOMPADDING", (0, 0), (-1, -1), pad),
            ]
            tbl.setStyle(TableStyle(ts))
            story.append(tbl)
            story.append(Spacer(1, 4 * mm * scale))
        elif btype == "image":
            # 图片块(公章等):worker 已把 source:'imgN' 解析成 url
            url = blk.get("url")
            if not url:
                continue
            wmm = float(blk.get("width_mm") or 38)
            if blk.get("overlay"):
                # 精确盖章:零高度叠加,压在上一行(需方盖章:)上。
                # 章**用实际尺寸,不随压页 scale 缩小**(否则压成一页时章会变得很小)。
                ov = _seal_overlay(url, wmm,
                                   float(blk.get("offset_x_mm") or 2),
                                   blk.get("offset_y_mm"), 1.0)
                if ov is not None:
                    story.append(ov)
            else:
                img = _image_flowable(url, wmm * scale)
                if img is not None:
                    story.append(img)
                    story.append(Spacer(1, 2 * mm * scale))

    if not story:
        story.append(Paragraph(" ", body))

    margin = max(8.0, 20 * scale)   # mm,下限 8mm
    lr = max(8.0, 18 * scale)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            topMargin=margin * mm, bottomMargin=margin * mm,
                            leftMargin=lr * mm, rightMargin=lr * mm)
    doc.build(story)
    return buf.getvalue()


def _page_count(data: bytes) -> int:
    try:
        from pypdf import PdfReader
        return len(PdfReader(io.BytesIO(data)).pages)
    except Exception:
        return 1


def _fetch_bytes(url: str) -> bytes | None:
    """下载图片(公章等)。Homebrew/容器 SSL 用 certifi 兜底。"""
    import ssl
    import urllib.request
    try:
        import certifi
        ctx = ssl.create_default_context(cafile=certifi.where())
    except Exception:
        ctx = ssl.create_default_context()
    try:
        return urllib.request.urlopen(url, context=ctx, timeout=20).read()
    except Exception:
        return None


def _seal_png(data: bytes) -> bytes:
    """把近白底像素抠成透明,让红章干净叠在文字上(手机拍的章多是白底)。失败原样返回。"""
    try:
        from PIL import Image as PILImage
        im = PILImage.open(io.BytesIO(data)).convert("RGBA")
        out = [(r, g, b, 0) if (r > 225 and g > 225 and b > 225) else (r, g, b, a)
               for (r, g, b, a) in im.getdata()]
        im.putdata(out)
        buf = io.BytesIO()
        im.save(buf, "PNG")
        return buf.getvalue()
    except Exception:
        return data


def _image_flowable(url: str, width_mm: float):
    """下载图片 → 抠白底 → reportlab Image flowable(等比缩放到 width_mm)。失败返回 None。"""
    from reportlab.lib.units import mm
    from reportlab.lib.utils import ImageReader
    from reportlab.platypus import Image as RLImage
    data = _fetch_bytes(url)
    if not data:
        return None
    png = _seal_png(data)
    try:
        iw, ih = ImageReader(io.BytesIO(png)).getSize()
        w = width_mm * mm
        h = w * ih / iw if iw else w
        img = RLImage(io.BytesIO(png), width=w, height=h)
        img.hAlign = "LEFT"   # 章靠左,落到「需方盖章」一侧,不居中飘中间
        return img
    except Exception:
        return None


def _seal_overlay(url: str, width_mm: float, offset_x_mm: float,
                  offset_y_mm, scale: float):
    """精确盖章:返回一个零高度 Flowable,把章画在自己流位置的上方左侧
    (压住上一行「需方盖章:」),不占版面、不推开后续内容,随排版自动跟位。"""
    from reportlab.lib.units import mm
    from reportlab.lib.utils import ImageReader
    from reportlab.platypus import Flowable
    data = _fetch_bytes(url)
    if not data:
        return None
    png = _seal_png(data)
    try:
        ir = ImageReader(io.BytesIO(png))
        iw, ih = ir.getSize()
        w = width_mm * mm
        h = w * ih / iw if iw else w
    except Exception:
        return None
    dx = offset_x_mm * mm
    # 默认让章竖向中心落在流线上(= 压在上一行签字行上,上下各探一半)
    dy = (float(offset_y_mm) * mm * scale) if offset_y_mm is not None else (-h / 2)

    class _Stamp(Flowable):
        def wrap(self, aw, ah):
            return (0, 0)          # 零尺寸,不挤占版面
        def draw(self):
            self.canv.drawImage(ir, dx, dy, width=w, height=h,
                                mask="auto", preserveAspectRatio=True)

    return _Stamp()


def build_pdf(spec: dict) -> bytes:
    """结构化文档 → .pdf bytes (中文用 STSong-Light, 无豆腐块)。
    spec: {filename, title?, blocks[], fit_pages?}
    blocks: [{type:'paragraph', text}, {type:'table', headers[], rows[][]}]
    fit_pages 设了(如 1)→ 自动等比缩小字号/边距,边渲染边数页,收敛到 ≤fit_pages 页(尽力)。
    不传 → 默认字号、自动分页(原行为)。
    """
    fit = spec.get("fit_pages")
    if not isinstance(fit, int) or fit < 1:
        return _build_pdf_bytes(spec, 1.0)
    last = None
    for scale in (1.0, 0.88, 0.78, 0.68, 0.6, 0.52, 0.45):
        data = _build_pdf_bytes(spec, scale)
        last = data
        if _page_count(data) <= fit:
            return data
    return last  # 缩到下限仍超页 → 返回最小那版(尽力)


# ── Claude tool schemas (dispatch 在 worker 层, 因需上传副作用) ────────────────

DOC_TOOL_SCHEMAS = [
    {
        "name": "make_excel",
        "description": "把结构化数据生成可下载的 Excel(.xlsx)文件。算完账/整理好表格后, 用户要下载表格时用。filename 用中文(无扩展名)。",
        "input_schema": {"type": "object", "properties": {
            "filename": {"type": "string", "description": "中文文件名, 无扩展名, 如 '订单核算表'"},
            "title": {"type": "string", "description": "表内大标题(可选)"},
            "sheet_name": {"type": "string"},
            "headers": {"type": "array", "items": {"type": "string"}},
            "rows": {"type": "array", "items": {"type": "array"}},
            "summary": {"type": "array", "items": {"type": "array"},
                        "description": "汇总行(可选), 会加粗"},
        }, "required": ["filename", "headers", "rows"]},
    },
    {
        "name": "make_pdf",
        "description": "把结构化文档生成可下载的 PDF 文件(中文正常显示)。报价单/合同/PI/正式文件用。filename 用中文(无扩展名)。"
                       "用户要「一页纸/压成一页/紧凑排版」时,传 fit_pages=1,系统会自动缩字号和边距把内容压进一页 —— 你**能**控制页数,别再说做不到。",
        "input_schema": {"type": "object", "properties": {
            "filename": {"type": "string", "description": "中文文件名, 无扩展名, 如 '报价单'"},
            "title": {"type": "string"},
            "blocks": {"type": "array", "items": {"type": "object"},
                       "description": "文档块: {type:'paragraph',text} 或 {type:'table',headers[],rows[][]} 或 "
                                      "{type:'image',overlay:true,width_mm:42}(盖公章:不用指定哪张图,系统自动挑用户传的方形红章、抠白底,"
                                      "overlay:true=精确压在上一行文字上 —— 把这块紧跟在「需方盖章:」那段之后即可;可选 offset_x_mm/offset_y_mm 微调)"},
            "fit_pages": {"type": "integer",
                          "description": "把内容压到几页内(用户要一页就传 1)。系统自动等比缩字号/边距适配。不传=默认字号自动分页。"},
        }, "required": ["filename", "blocks"]},
    },
]

# 文件类型 → (build 函数, 扩展名, mime)
DOC_BUILDERS = {
    "make_excel": (build_excel, "xlsx",
                   "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
    "make_pdf": (build_pdf, "pdf", "application/pdf"),
}
