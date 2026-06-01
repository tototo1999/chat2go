"""chat2go.ai Modal worker — 文本对话 serverless 替代 Hermes daemon。

调用链:
    chat.html 用户发消息
        → INSERT messages (role=user/expert)
        → fetch Edge Function chat2go-ingest
            → INSERT placeholder AI 消息 (content='...')
            → POST 本 worker (fire-and-forget)
                → 本 worker 拉历史 → 调 Claude → UPDATE placeholder
                    → Supabase Realtime → chat.html 自动重渲

部署(2026-05-25):
    # 复用 speak2go-secrets(已有 SUPABASE_URL / SUPABASE_SERVICE_ROLE_KEY)
    modal secret create chat2go-extras ANTHROPIC_API_KEY=sk-ant-...
    modal deploy chat2go_worker.py
    # 拿到 endpoint,塞到 Supabase Edge Function chat2go-ingest 的 CHAT2GO_MODAL_WORKER_URL
    # Worker 端目前不做 token 校验,Edge Function 端 CHAT2GO_MODAL_WORKER_TOKEN
    # 仅作 header 透传(规避 Modal endpoint 公开 URL 被滥用是后续 TODO)。

成本预估:
    Claude Sonnet 4.6 单轮 ~2-4k input + 1-2k output ≈ $0.015-0.03/轮
    单房日均 50 轮 × 30 房 ≈ $20-45/日,低于 Hermes mini 电费 + 维护成本
"""
from __future__ import annotations

import hmac
import json
import os
import time
from typing import Any

import modal

# FastAPI Request 类型: Modal 容器靠这个注解才会注入 Request 对象(读 header);
# 本地测试环境无 fastapi → 降级为 object, 仅保证模块可 import(测试不调 ingest)。
try:
    from fastapi import Request as _FastAPIRequest
except ImportError:  # pragma: no cover
    _FastAPIRequest = object

import trade_accounting as ta  # 外贸会计核算工具 (子项目②)
import doc_gen as dg            # Excel/PDF 生成 (子项目③)

# Supabase Storage bucket (public, 已存在)
STORAGE_BUCKET = "chat-uploads"

# ── Modal app + image ────────────────────────────────────────────────────────
app = modal.App("chat2go-worker")

image = (
    modal.Image.debian_slim(python_version="3.11")
    .pip_install(
        "supabase>=2.10.0",
        "anthropic>=0.39.0",
        "httpx>=0.27",
        "fastapi[standard]>=0.115",
        "openpyxl>=3.1",      # Excel 生成(③)+ 读取(读文件)
        "reportlab>=4.0",     # PDF 生成 + 内置中文 CID 字体 (子项目③)
        "pypdf>=4.0",         # PDF 文本提取 (读文件)
        "python-docx>=1.1",   # Word 文本提取 (读文件)
    )
    # 本地纯 python 模块(会计工具②, 文档生成③), 随 image 带进容器
    .add_local_python_source("trade_accounting")
    .add_local_python_source("doc_gen")
)

# 复用 speak2go-secrets(已有 SUPABASE_URL / SUPABASE_SERVICE_ROLE_KEY) + chat2go-extras
# (只装 ANTHROPIC_API_KEY)。这样不用把 service_role 复制成两份,也不用从
# Supabase dashboard 重新取。
secrets = [
    modal.Secret.from_name("speak2go-secrets"),
    modal.Secret.from_name("chat2go-extras"),
]

# ── 行业 system prompt(从 supabase/functions/chat-ai/index.ts 同步过来,加 命理) ────

INDUSTRY_PROMPTS: dict[str, str] = {
    "外贸": """你是一位外贸行业的 AI 助手,专注于帮助跟单员处理外贸业务。
你熟悉:合同生成(FOB/CIF/CFR 条款)、信用证、提单、装箱单、报关、物流跟踪、汇率结算。
用简洁、专业的中文回复。对于合同条款,直接给出可复用的模板片段。""",

    "健身": """你是一位健身行业的 AI 助手,帮助健身教练管理学员和课程。
你熟悉:学员 CRM 记录、训练计划制定、体测数据分析、课程排期、营养建议。
用鼓励、专业的语气回复,给出具体可操作的建议。""",

    "地产": """你是一位房地产行业的 AI 助手,帮助中介提升带客效率。
你熟悉:房源分析、周边配套研报、客户意向判断、谈判策略、合同要点。
回复简洁有力,优先给出数据支撑的分析。""",

    "教育": """你是一位教育行业的 AI 助手,帮助学生和教师提升学习效果。
你熟悉:课件整理、知识点讲解、习题解析、学习计划制定、考点总结。
用清晰、有条理的方式解释,适合中学生和大学生理解。""",

    "量化": """你是一位量化交易领域的 AI 助手,帮助用户理解量化策略。
你熟悉:回测逻辑、因子分析、风险控制、仓位管理、Python/pandas 数据处理。
面向非专业用户时,用类比和简单语言解释复杂概念。""",

    "医疗": """你是一位医疗辅助 AI 助手,帮助医生处理患者咨询和诊疗记录整理。
重要:你只提供信息参考,不做最终诊断,每次回复都提醒用户以医生判断为准。
你熟悉:常见症状描述规范、病历整理格式、患者沟通话术。""",

    "算命": """你是一位命理研习 AI 助手,辅助命理师与求测者沟通。
你熟悉:八字四柱、紫微斗数、奇门遁甲、易经卦象、风水基础。
**重要边界**:命理输出只作传统文化研习与心理引导参考,不替代医疗、法律、金融的专业建议;
不预测寿数、不下生死断语、不鼓动改命消业的高额服务。
回复风格:中文,先列出基本盘面/卦象/格局,再给读法,最后给一两句生活化的建议;
用户给生辰八字时按公历转干支后排盘,看不全的字段直接问追问。""",

    "康复治疗": """你是一位康复治疗 AI 助手,辅助康复治疗师与就诊者沟通(Well2GO.ai 平台)。
你熟悉:首次评估(主诉/疼痛/活动度)、功能筛查(肌力/平衡/步态)、体态分析、
治疗方案设计(手法/运动/物理因子)、训练打卡、复诊评估、宣教与预防、结案与转归。
**重要边界**:你只作专业辅助,不下临床诊断、不开处方、不替代影像/医师判断;
涉及急性外伤、神经科红旗征、剧烈疼痛 → 提醒就医,不擅自给方案。
回复风格:简体中文,**先听症再问关键、再给建议**,优先用 SOAP 或 OPQRSTU 框架,
方案要可执行(动作名+组数+次数+频率),配上"何时停止/进阶"提示。""",

    "英语口语": """You are a 1-on-1 English speaking coach for Speak2GO.ai. Voice: NYC casual, warm, sharp.

You handle three modes seamlessly:
- **Pronunciation drills**: short corrections, **romanized phonetics** (e.g. `SNEIK` for snake, `e·KOM·e·deit` for accommodate — lowercase ASCII, stressed syllable in CAPS, `·` between syllables; **never use IPA symbols**), model sentence + student repeats.
- **Conversational practice**: pick a context (cafe, job interview, small talk), play one role, throw realistic curveballs.
- **Knowledge & test prep**: explain grammar/vocab/idioms with 1-line rule + 2 vivid examples.

Hard rules:
- **Always answer in English** even if the student writes Chinese — translate intent, reply in English, only use Chinese as a glossary aside when blocking comprehension.
- Never call yourself "Chat2Go" or any other platform name. You are part of Speak2GO.ai / "our class".
- Keep replies tight: 2-4 sentences unless the student asks for a deep dive.
- Praise honestly, correct directly. NYC casual means real talk, not flowery.""",

    "英文作文": """你是一位英语写作批改老师(Essay2GO 系统)。学生会把英文作文直接发在聊天里(可能附上题目),你的任务是**按五维 rubric 批改并给出可执行的提升建议**。

只批改学生**最新发来的那篇作文**(忽略历史聊天里的旧作文)。如果学生发的不是作文(只是提问/打招呼),就正常用中文简短回应、引导他写,不要硬批改。

**输出必须是 markdown**,严格按下面结构:

## 📊 评分
- 内容 Content: X/20 — 一句点评
- 结构 Structure: X/20 — 一句点评
- 语法 Grammar: X/20 — 一句点评
- 用词 Vocabulary: X/20 — 一句点评
- 衔接 Cohesion: X/20 — 一句点评
- **总分 X/100 · 等级 <CEFR 如 B1/B2/C1>**

## ✏️ 逐句/逐段修改
(挑 3-6 处最值得改的:`原句` → 问题(中文一句) → ✅ 改后。只挑高价值的,不要逐字挑刺。)

## 🌟 高分改写示范
(选作文里最弱的一段,给出一段地道的高分改写,让学生看到差距。)

## 💡 下一步
(1-2 条最该练的方向,中文。)

批改完,在**最后另起一行**输出这一行机器可读的进度标记(给系统记录写作进度用,学生也能看到):
```essay-score
{"score": <总分整数>, "band": "<CEFR>", "date": "<YYYY-MM-DD>"}
```

语气:鼓励但直接,先肯定亮点再指问题。点评用中文,改写/示范用英文。""",
}

# 同义词映射(数据库里历史写法对齐 INDUSTRY_PROMPTS 的 key)
INDUSTRY_ALIASES = {
    "命理": "算命",
    "外贸跟单": "外贸",   # tradego 老房复用外贸 prompt
}

# 外贸会计核算指引 — 对任何外贸房追加到 system prompt(不管默认还是大咖自定义 prompt),
# 让 AI 知道有精确计算工具且必须调用(子项目②, 重建 Hermes 时代 contract_lib/excel_lib 能力)。
TRADE_ACCOUNTING_GUIDE = """

## 会计核算(重要)
你配有一套**精确计算工具**,任何涉及金额的核算都**必须调用工具**算,绝不自己心算口算:
- calc_unit_cost — 单位成本(采购+运费+关税+杂费)/数量
- quote_from_margin — 按目标利润率倒推报价 + FOB/CIF/CFR 换算
- order_pnl — 单笔订单损益(毛利/净利/利润率)
- fx_convert — 汇率换算(汇率由用户给出, 你不要编汇率)
- export_rebate — 出口退税额
- commission — 佣金核算 + 净利还原
- reconcile — 对账 + 账期 aging
规则:利润率/税率/退税率传**小数**(20% 传 0.2)。缺关键数字(数量、汇率、税率等)就先**追问**, 不要假设。
算完用 **markdown 表格**清晰展示结果和明细(工具返回的 breakdown),让用户一眼看懂每一步。

## 生成文件(合同 / PI / CI / 装箱单 / 报价单 / Excel)— 铁律
你配有 **make_pdf**(生成 PDF)和 **make_excel**(生成 Excel)工具。要给用户任何可下载文件时:
1. **必须真的调用工具生成**。用户要「全套单证」就为**每一份**各调一次 make_pdf(PI 一次、CI 一次、装箱单一次…),一份文件 = 一次工具调用。
2. **绝对禁止自己在正文里写下载链接 / URL / 文件路径**(例如 `gen/xxx.pdf`、`[合同](http://…/gen/…pdf)`)。这些路径你**编不出来也不许编** —— 真实链接只能由工具返回。生成的文件会**自动作为可下载附件卡片显示在本条消息下方**,你只需用文字简述生成了哪几份即可,**不要写任何 markdown 链接**。
3. **没调用工具,就绝对不能说「已生成」「点击下载」**。没生成成功就如实说,别谎称已生成。"""


def _is_trade_room(industry: str) -> bool:
    """该房是否外贸房(挂会计工具)。基于 industry, 不受 room.system_prompt 覆盖影响。"""
    return INDUSTRY_ALIASES.get(industry, industry) == "外贸"

DEFAULT_SYSTEM = """你是 Chat2GO.ai 平台的专属 AI 助手,工作在"Chat 调试室"中。
你的目标是帮助小白和大咖共同理清需求,展示 AI 能做什么,最终为小白交付一个可以独立使用的专属 AI 助手。
请用简洁、专业的中文回复。如果需要更多信息才能准确回答,请直接追问关键细节。"""

# ── Helpers ───────────────────────────────────────────────────────────────────


def _sb_client():
    """Supabase service-role client(绕过 RLS,读全房消息)。"""
    from supabase import create_client
    return create_client(
        os.environ["SUPABASE_URL"],
        os.environ["SUPABASE_SERVICE_ROLE_KEY"],
    )


def _anthropic_client():
    from anthropic import Anthropic
    return Anthropic(api_key=os.environ["ANTHROPIC_API_KEY"])


# 房间内最近 N 轮上下文 (足够保留 token 上下文,又不爆 200k)
HISTORY_LIMIT = 40
# Claude 模型默认(rooms.model 为空时用)
DEFAULT_MODEL = "claude-sonnet-4-6"
# 单次回复 max_tokens
MAX_TOKENS = 4096
# 外贸 tool-use 单轮最多工具调用次数(防失控/防成本爆炸)
MAX_TOOL_ITERS = 5


def _load_history(sb, room_id: str, channel: str, before_message_id: str) -> list[dict]:
    """拉本房本频道 N 条历史消息,按时间升序返回。
    placeholder 自身(role=ai content='...') 也会被拉到,要在这一步过滤掉。
    """
    # 拿 trigger 消息的 created_at,确保上下文截止到它(不含 placeholder)。
    # 用 maybe_single: 触发消息偶发查到 0 行(并发/读时序)时返回 data=None 而不抛
    # PGRST116, 退化成「不设截止、照常拉历史」, 而不是整条回复崩成「AI 调用失败」。
    trig = sb.table("messages").select("created_at").eq("id", before_message_id).maybe_single().execute()
    cutoff = trig.data["created_at"] if trig and trig.data else None

    q = sb.table("messages") \
        .select("id, role, content, type, attachments, created_at, channel") \
        .eq("room_id", room_id) \
        .eq("channel", channel) \
        .order("created_at", desc=True) \
        .limit(HISTORY_LIMIT)
    if cutoff:
        q = q.lte("created_at", cutoff)
    rows = q.execute().data or []
    rows.reverse()  # 升序
    # 防御性过滤:placeholder 占位 AI 消息("..." / 以 ⚠️ 开头) 不该进上下文
    cleaned = [r for r in rows if not _is_placeholder(r)]
    return cleaned


def _is_placeholder(row: dict) -> bool:
    if row.get("role") != "ai":
        return False
    c = (row.get("content") or "").strip()
    return c == "..." or c.startswith("⚠️")


# Claude Vision 支持的图片格式
_VISION_EXTS = (".png", ".jpg", ".jpeg", ".gif", ".webp")


def _is_safe_image_url(url: str) -> bool:
    """SSRF 防护(审计#5): 只允许本项目 Supabase Storage 的 https URL 传给 Claude Vision。
    Anthropic 服务端会 fetch image url-source, 不限制 = 任意外联/内网探测。"""
    if not isinstance(url, str) or not url.startswith("https://"):
        return False
    base = (os.environ.get("SUPABASE_URL") or "").rstrip("/")
    if not base:
        return False
    return url.startswith(base + "/storage/v1/object/")


def _image_atts(row: dict) -> list[dict]:
    """从一条消息的 attachments 里挑出图片附件(按 mime 或扩展名), 且 URL 通过 SSRF 校验。"""
    atts = row.get("attachments")
    if not isinstance(atts, list):
        return []
    imgs = []
    for a in atts:
        if not isinstance(a, dict) or not a.get("url"):
            continue
        if not _is_safe_image_url(a["url"]):
            continue  # 拒绝非本项目 Storage 的 URL(SSRF)
        mime = (a.get("mime_type") or "").lower()
        name = (a.get("name") or "").lower()
        if mime.startswith("image/") or name.endswith(_VISION_EXTS):
            imgs.append(a)
    return imgs


# ── 读文件: 非图片附件文本提取(重建 bridge.py 功能#7, cutover 后丢失)─────────
# 文档附件(PDF/Excel/Word/文本)→ 从 Storage 下载 → 抽文本 → 注入上下文让 AI 能读。
_DOC_EXTS = (".pdf", ".xlsx", ".xlsm", ".docx", ".txt", ".csv",
             ".md", ".json", ".xml", ".html", ".log", ".tsv")
MAX_DOC_BYTES = 8_000_000   # 单附件最多下载 8MB
MAX_DOC_CHARS = 16_000      # 单附件注入文本上限(防爆 token)


def _storage_path_from_url(url: str) -> str | None:
    """从本项目 Storage URL 解析对象 path(.../object/.../chat-uploads/<path>)。
    非本项目 Storage 的 URL → None(SSRF: 只下自己桶里的文件)。"""
    if not isinstance(url, str):
        return None
    base = (os.environ.get("SUPABASE_URL") or "").rstrip("/")
    if not base or not url.startswith(base + "/storage/v1/object/"):
        return None
    marker = "/" + STORAGE_BUCKET + "/"
    i = url.find(marker)
    if i < 0:
        return None
    path = url[i + len(marker):].split("?", 1)[0]
    return path or None


def _doc_atts(row: dict) -> list[dict]:
    """挑出可读文本的文档附件(非图片, 且有本项目 storage path)。返回 [{name, path}]。"""
    atts = row.get("attachments")
    if not isinstance(atts, list):
        return []
    out = []
    for a in atts:
        if not isinstance(a, dict):
            continue
        name = (a.get("name") or "").lower()
        mime = (a.get("mime_type") or "").lower()
        if mime.startswith("image/") or name.endswith(_VISION_EXTS):
            continue  # 图片走 Vision, 不在这里
        if not name.endswith(_DOC_EXTS):
            continue
        path = a.get("storage_path") or _storage_path_from_url(a.get("url") or "")
        if not path:
            continue
        out.append({"name": a.get("name") or "文件", "path": path})
    return out


def _download_storage(sb, path: str) -> bytes | None:
    try:
        data = sb.storage.from_(STORAGE_BUCKET).download(path)
    except Exception:
        return None
    return data[:MAX_DOC_BYTES] if data else None


def _extract_doc_text(name: str, data: bytes) -> str:
    """按扩展名抽文本。不支持/解析失败 → 返回空或提示, 绝不抛。"""
    import io
    n = (name or "").lower()
    try:
        if n.endswith(".pdf"):
            from pypdf import PdfReader
            r = PdfReader(io.BytesIO(data))
            return "\n".join((p.extract_text() or "") for p in r.pages).strip()
        if n.endswith((".xlsx", ".xlsm")):
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            parts = []
            for ws in wb.worksheets:
                parts.append(f"# {ws.title}")
                for rw in ws.iter_rows(values_only=True):
                    cells = [str(c) for c in rw if c is not None]
                    if cells:
                        parts.append("\t".join(cells))
            return "\n".join(parts).strip()
        if n.endswith(".docx"):
            from docx import Document
            d = Document(io.BytesIO(data))
            return "\n".join(p.text for p in d.paragraphs).strip()
        if n.endswith((".txt", ".csv", ".md", ".json", ".xml", ".html", ".log", ".tsv")):
            return data.decode("utf-8", "replace").strip()
    except Exception as e:  # noqa: BLE001
        return f"(无法解析 {name}: {type(e).__name__})"
    return ""


def _doc_text_block(sb, docs: list[dict]) -> str:
    """下载并抽取一条消息的所有文档附件文本, 拼成「附件内容」段(供 LLM 读)。"""
    if sb is None or not docs:
        return ""
    chunks = []
    for d in docs:
        data = _download_storage(sb, d["path"])
        if not data:
            chunks.append(f"[附件 {d['name']}: 下载失败]")
            continue
        text = _extract_doc_text(d["name"], data)
        if not text:
            chunks.append(f"[附件 {d['name']}: 无可提取文本(可能是扫描件, 请截图发我走读图)]")
            continue
        if len(text) > MAX_DOC_CHARS:
            text = text[:MAX_DOC_CHARS] + f"\n…(文件较长, 已截断到 {MAX_DOC_CHARS} 字)"
        chunks.append(f"[附件 {d['name']} 内容如下]\n{text}")
    return "\n\n".join(chunks)


def _to_blocks(content) -> list[dict]:
    """字符串 content → [{type:text}] block 列表(已是 list 则原样返回)。"""
    if isinstance(content, list):
        return content
    return [{"type": "text", "text": content}] if content else []


def _build_messages(history: list[dict], sb=None) -> list[dict]:
    """history → Anthropic messages array。
    role 映射:user / expert → 'user';ai → 'assistant'。
    连续 same role 合并(纯文本拼字符串;含图片转 content block 列表)。
    图片附件 → Claude Vision image block(url source),让 AI 能读图/OCR。
    文档附件(PDF/Excel/Word/文本)→ 下载抽文本拼进消息,让 AI 能读文件(需传 sb)。
    """
    out: list[dict] = []
    for r in history:
        role = "assistant" if r.get("role") == "ai" else "user"
        text = (r.get("content") or "").strip()
        imgs = _image_atts(r) if role == "user" else []  # 只对用户侧消息读图
        docs = _doc_atts(r) if (role == "user" and sb is not None) else []
        if not text and not imgs and not docs:
            continue
        # 把 user/expert 区分塞进 text 前缀,让 LLM 看到角色差异
        if text:
            if r.get("role") == "expert":
                text = f"[大咖] {text}"
            elif r.get("role") == "user":
                text = f"[小白] {text}"
        # 文档附件 → 抽文本拼到消息正文(让 AI 读文件)
        if docs:
            doc_text = _doc_text_block(sb, docs)
            if doc_text:
                text = (text + "\n\n" + doc_text) if text else doc_text

        if imgs:
            # 含图片 → content block 列表: 图片在前, 文字在后
            blocks: list[dict] = [
                {"type": "image", "source": {"type": "url", "url": a["url"]}}
                for a in imgs
            ]
            if text:
                blocks.append({"type": "text", "text": text})
            cur_content: Any = blocks
        else:
            cur_content = text

        if out and out[-1]["role"] == role:
            prev = out[-1]["content"]
            if isinstance(prev, str) and isinstance(cur_content, str):
                out[-1]["content"] = prev + "\n\n" + cur_content
            else:
                out[-1]["content"] = _to_blocks(prev) + _to_blocks(cur_content)
        else:
            out.append({"role": role, "content": cur_content})
    # 必须以 user 起头
    while out and out[0]["role"] != "user":
        out.pop(0)
    return out


def _resolve_system_prompt(industry: str, room_system_prompt: str) -> str:
    """room.system_prompt 非空时优先(大咖可自定义),否则查 INDUSTRY_PROMPTS,再否则 default。"""
    if room_system_prompt and room_system_prompt.strip():
        return room_system_prompt.strip()
    key = INDUSTRY_ALIASES.get(industry, industry)
    return INDUSTRY_PROMPTS.get(key, DEFAULT_SYSTEM)


# room-scope memory 注入上限(防止 system prompt 过长)
MEMORY_LIMIT = 30


def _load_memories(sb, room_id: str, limit: int = MEMORY_LIMIT) -> list[dict]:
    """SELECT 本房 scope='room' 的 memories,按时间升序返回最近 N 条。
    迁移自 Hermes 老 sync_memory 机制,让大咖的偏好/事实沉淀跨对话保留。
    """
    rows = sb.table("memories") \
        .select("content, tags, created_at") \
        .eq("scope", "room") \
        .eq("scope_id", room_id) \
        .order("created_at", desc=True) \
        .limit(limit) \
        .execute().data or []
    rows.reverse()
    return rows


def _format_memories(memories: list[dict]) -> str:
    """把 memory 行拼成 markdown 段,前置到 system prompt。空时返回空串。"""
    if not memories:
        return ""
    lines = ["", "## 大咖偏好与历史沉淀(按时间顺序,请遵守)"]
    for m in memories:
        tag_str = " · ".join(m.get("tags") or [])
        prefix = f"[{tag_str}] " if tag_str else ""
        lines.append(f"- {prefix}{(m.get('content') or '').strip()}")
    return "\n".join(lines) + "\n"


def _looks_markdown(text: str) -> bool:
    """简单启发:含 ## / **/ - 列表 / ```代码块 等 → markdown 类型。"""
    triggers = ("##", "**", "```", "\n- ", "\n* ", "\n1. ")
    return any(t in text for t in triggers)


def _extract_text(resp) -> str:
    out = ""
    for blk in resp.content:
        if getattr(blk, "type", None) == "text":
            out += getattr(blk, "text", "")
    return out


def _make_doc_and_upload(sb, tool_name: str, tool_input: dict) -> tuple[dict, dict]:
    """生成 Excel/PDF → 上传 Supabase Storage → 返回 (attachment, tool_result)。
    storage 路径 ASCII(uuid), 中文名只放 attachment.name(中文路径 Storage 会 400)。"""
    builder, ext, mime = dg.DOC_BUILDERS[tool_name]
    spec = tool_input or {}
    data = builder(spec)
    display_name = (spec.get("filename") or "文件") + "." + ext
    path = dg.storage_path(spec.get("filename") or "file", ext)
    sb.storage.from_(STORAGE_BUCKET).upload(
        path, data, {"content-type": mime, "upsert": "true"},
    )
    url = sb.storage.from_(STORAGE_BUCKET).get_public_url(path)
    attachment = {
        "name": display_name, "url": url, "size": len(data),
        "mime_type": mime, "storage_path": path,
    }
    tool_result = {"ok": True, "url": url, "name": display_name, "size_bytes": len(data),
                   "note": "文件已生成并附在本条消息下,用户可直接下载"}
    return attachment, tool_result


def _run_completion(cli, sb, model: str, system: str, messages: list[dict],
                    is_trade: bool) -> tuple[str, list[dict]]:
    """返回 (最终文字, 生成的文件 attachments)。
    非外贸房:单次调用。外贸房:带会计+文档工具的 tool-use 循环(最多 MAX_TOOL_ITERS 次)。"""
    attachments: list[dict] = []
    if not is_trade:
        resp = cli.messages.create(
            model=model, max_tokens=MAX_TOKENS, system=system, messages=messages,
        )
        return _extract_text(resp).strip() or "(AI 没有返回内容)", attachments

    tools = ta.TOOL_SCHEMAS + dg.DOC_TOOL_SCHEMAS
    convo = list(messages)
    for _ in range(MAX_TOOL_ITERS):
        resp = cli.messages.create(
            model=model, max_tokens=MAX_TOKENS, system=system,
            messages=convo, tools=tools,
        )
        tool_uses = [b for b in resp.content if getattr(b, "type", None) == "tool_use"]
        if not tool_uses:
            return _extract_text(resp).strip() or "(AI 没有返回内容)", attachments
        # 回放 assistant 的 tool_use turn, 再附上每个 tool_result
        convo.append({"role": "assistant", "content": resp.content})
        results = []
        for tu in tool_uses:
            if tu.name in dg.DOC_BUILDERS:
                try:
                    att, out = _make_doc_and_upload(sb, tu.name, tu.input or {})
                    attachments.append(att)
                except Exception as e:  # noqa: BLE001
                    out = {"error": f"文件生成/上传失败: {type(e).__name__}: {e}"}
            else:
                out = ta.dispatch(tu.name, tu.input or {})
            results.append({
                "type": "tool_result",
                "tool_use_id": tu.id,
                "content": json.dumps(out, ensure_ascii=False),
            })
        convo.append({"role": "user", "content": results})
    # 用尽迭代次数:再要一次最终文字总结(不给 tools, 逼它收尾)
    resp = cli.messages.create(
        model=model, max_tokens=MAX_TOKENS, system=system, messages=convo,
    )
    return _extract_text(resp).strip() or "(AI 没有返回内容)", attachments


def _check_worker_auth(authorization: str) -> bool:
    """审计#1: env-gated bearer token 校验。
    CHAT2GO_MODAL_WORKER_TOKEN 未设 → 放行(灰度: 部署不破线, 两边密钥配齐前不强制)。
    已设 → 必须 Authorization: Bearer <token> 常数时间匹配, 否则拒。"""
    expected = os.environ.get("CHAT2GO_MODAL_WORKER_TOKEN", "")
    if not expected:
        return True
    got = (authorization or "")
    if not got.startswith("Bearer "):
        return False
    return hmac.compare_digest(got[len("Bearer "):].strip(), expected)


def _verify_placeholder(sb, placeholder_id: str, room_id: str) -> bool:
    """审计#1 纵深: 只允许更新本房 role='ai' 的占位消息, 防止攻击者传任意 message_id
    覆盖用户/大咖正文(worker 持 service-role 绕 RLS)。"""
    try:
        row = sb.table("messages").select("role, room_id").eq("id", placeholder_id) \
                .single().execute().data
    except Exception:
        return False
    return bool(row) and row.get("role") == "ai" and row.get("room_id") == room_id


def _update_placeholder(sb, placeholder_id: str, content: str, msg_type: str | None = None,
                        attachments: list[dict] | None = None) -> None:
    payload: dict[str, Any] = {"content": content}
    if msg_type:
        payload["type"] = msg_type
    if attachments:
        payload["attachments"] = attachments
    sb.table("messages").update(payload).eq("id", placeholder_id).execute()


# ── Modal entry point ─────────────────────────────────────────────────────────

@app.function(
    image=image,
    secrets=secrets,
    timeout=600,   # 10min — 单轮 Claude 极少超 2 min,留余量
    cpu=1,
    memory=1024,
)
@modal.fastapi_endpoint(method="POST")
def ingest(payload: dict, request: _FastAPIRequest) -> dict:
    """Modal web endpoint,被 supabase/functions/chat2go-ingest 调用。
    request 必须带 _FastAPIRequest 注解, FastAPI 才会注入 Request 对象(读 header)。"""
    # 审计#1: 鉴权(env-gated). 公开 URL, 防白嫖 Claude/Storage + 跨房篡改。
    from fastapi.responses import JSONResponse
    if not _check_worker_auth(request.headers.get("authorization", "")):
        return JSONResponse({"ok": False, "error": "unauthorized"}, status_code=401)

    placeholder_id = payload.get("placeholder_id")
    room_id = payload.get("room_id")
    trigger_message_id = payload.get("trigger_message_id")
    channel = payload.get("channel", "main")
    industry = payload.get("industry", "")
    room_system_prompt = payload.get("system_prompt", "") or ""
    model = (payload.get("model") or "").strip() or DEFAULT_MODEL

    if not (placeholder_id and room_id and trigger_message_id):
        return {"ok": False, "error": "missing_required_fields"}

    sb = _sb_client()
    # 审计#1 纵深: 只更新本房 role='ai' 占位, 防覆盖任意消息正文
    if not _verify_placeholder(sb, placeholder_id, room_id):
        return JSONResponse({"ok": False, "error": "invalid_placeholder"}, status_code=403)
    t0 = time.time()

    try:
        history = _load_history(sb, room_id, channel, trigger_message_id)
        messages = _build_messages(history, sb)
        if not messages:
            _update_placeholder(sb, placeholder_id,
                                "⚠️ 没拉到上下文消息,无法回复。请重新发一条试试。")
            return {"ok": False, "error": "empty_history"}

        base_system = _resolve_system_prompt(industry, room_system_prompt)
        mem_rows = _load_memories(sb, room_id)
        system = base_system + _format_memories(mem_rows)

        # 外贸房:追加会计核算指引 + 挂计算工具(不受 room.system_prompt 覆盖影响)
        is_trade = _is_trade_room(industry)
        if is_trade:
            system = system + TRADE_ACCOUNTING_GUIDE

        cli = _anthropic_client()
        out_text, doc_attachments = _run_completion(cli, sb, model, system, messages, is_trade)

        msg_type = "markdown" if _looks_markdown(out_text) else "text"
        _update_placeholder(sb, placeholder_id, out_text, msg_type=msg_type,
                            attachments=doc_attachments or None)

        dt = time.time() - t0
        return {
            "ok": True,
            "placeholder_id": placeholder_id,
            "elapsed_s": round(dt, 1),
            "model": model,
            "input_msgs": len(messages),
            "memories": len(mem_rows),
            "output_chars": len(out_text),
        }
    except Exception as e:
        print(f"[error] chat2go ingest failed: {e!r}")
        try:
            _update_placeholder(sb, placeholder_id,
                                f"⚠️ AI 调用失败:{type(e).__name__} {str(e)[:160]}")
        except Exception:
            pass
        return {"ok": False, "error": str(e), "placeholder_id": placeholder_id}


if __name__ == "__main__":
    print("chat2go_worker.py — Modal app skeleton OK")
    print("Deploy: modal deploy chat2go_worker.py")
    print("Endpoint will be: https://<workspace>--chat2go-worker-ingest.modal.run")
